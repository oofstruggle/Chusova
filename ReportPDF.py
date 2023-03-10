import csv
from datetime import datetime
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import numpy as np
import matplotlib.pyplot as plt
from jinja2 import Environment, FileSystemLoader
import pdfkit
import cProfile
import dateparser
import datetime as dt
import csv_file_separator as csv_files_generator


currency_to_rub = {'AZN': 35.68, 'BYR': 23.91, 'EUR': 59.90, 'GEL': 21.74, 'KGS': 0.76, 'KZT': 0.13, 'RUR': 1,
                   'UAH': 1.64, 'USD': 60.66, 'UZS': 0.0055
                   }


class UsersInput:
    """Класс для пользовательского ввода по шаблону"""
    def __init__(self):
        """Инициализация объекта UsersInput"""
        self.compiled_file = input('Введите название файла: ')
        self.position_title = input('Введите название профессии: ')
        self.compiled_file = self.check_file_name(self.compiled_file)
        self.position_title = self.validate_position_name(self.position_title)

    @staticmethod
    def validate_file_name(compiled_file):
        """
        Валидация названия файла
        Args:
            compiled_file (str): Выбранный файл
        """
        if compiled_file == '' or '.' not in compiled_file:
            print('Некорректное название файла')
            exit()
        return compiled_file

    @staticmethod
    def validate_position_name(position_title):
        """Валидация названия профессии

        Args:
            position_title (str): Выбранное название профессии

        Returns:
            str: корректное название профессии
        """
        if position_title == '':
            print('Некорректное название профессии')
            exit()
        return position_title


class DataSet:
    """Класс, представляющий сбор данных"""
    def __init__(self, compiled_file):
        """Инициализация объекта DataSet

        Args:
            compiled_file (str): Выбранный файл

        >>> type(DataSet("vacancies_by_year.csv", "Аналитик")).__name__
        'DataSet'
        >>> DataSet("vacancies_by_year.csv", "Аналитик").profession
        'Аналитик'
        >>> DataSet("vacancies_by_year.csv", "Аналитик").file_name
        'vacancies.csv'
        """
        self.reader = []
        for row in csv.reader(open(compiled_file, encoding='utf_8_sig')):
            self.reader += [row]
        if len(self.reader) == 0:
            print('Пустой файл')
            exit()
        self.columns_names = self.reader[0]
        self.vacancies_data = []
        for row in self.reader[1:]:
            if len(row) == len(self.columns_names) and row.count('') == 0:
                self.vacancies_data += [row]
        if len(self.vacancies_data) == 0:
            print('Нет данных')
            exit()

    def parse_csv(self):
        """Считывает данные из csv-файла и форматирует их"""

        csv_strings, titles = reader.csv_reader(self.file_name)
        csv_files_generator.parse_by_years(csv_strings, titles)
        self.vacancies_objects = reader.csv_filer(csv_strings, titles, self.create_vacancy)

class Vacancy:
    """Класс, представляющий вакансии"""

    name: str
    salary_from: int or float
    salary_to: int or float
    salary_currency: str
    area_name: str
    published_at: str
    salary: str

    def __init__(self, vacancy):
        """Инициализация объекта Vacancy

            Args:
                vacancy (str): вакансия

        >>> Vacancy({"name": "Аналитик", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Москва",  "published_at": "2022-07-05T18:19:30+0300"}).name
        'Аналитик'
        >>> Vacancy({"name": "Аналитик", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Москва",  "published_at": "2022-07-05T18:19:30+0300"}).salary
        55000.0
        >>> Vacancy({"name": "Аналитик", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Москва",  "published_at": "2022-07-05T18:19:30+0300"}).published_at
        2022
        >>> Vacancy({"name": "Аналитик", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Москва",  "published_at": "2022-07-05T18:19:30+0300"}).area_name
        'Москва'
        >>> type(Vacancy({"name": "Аналитик", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Москва",  "published_at": "2022-07-05T18:19:30+0300"})).__name__
        'Vacancy'
        """
        for key, value in vacancy.items():
            self.__setattr__(key, self.formatter(key, value))

    @staticmethod
    def formatter(key, value):
        """Форматирование значения в float или int в зависимости от данных

            Args:
                key (str): ключ словаря
                value (str or int or float): значение по ключу

            Returns:
                float or int: сконвертированное число
        """

        if key in ['salary_from', 'salary_to']:
            return float(value)
        if key == 'published_at':
            return int(datetime.strptime(value, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))
        return value


class Salary:
    """Класс представления зарплаты"""
    def __init__(self, salary_from, salary_to, salary_currency):
        """Инициализация объекта Salary

            Args:
                salary_from (str or int or float): нижняя граница оклада
                salary_to (str or int or float): верхняя граница оклада
                salary_currency (str): требуемая валюта
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency


class SalaryDict:
    """Класс формирования словаря для подсчёта статистики зарплат"""
    def __init__(self):
        """Инициализация объекта SalaryDict"""
        self.salary_dict = {}
        self.__average_salary_dict = {}

    def append_salary(self, key, salary):
        """Добавить зарплату в словарь

            Args:
                key: ключ словаря
                salary: зарплата

            Returns:
                array: список зарплат по ключу
        """

        if self.salary_dict.get(key) is None:
            self.salary_dict[key] = []
        return self.salary_dict[key].append(salary)

    def calculate_average_salary(self):
        """Подсчёт средней цифры заработной платы"""
        for key, value in self.salary_dict.items():
            self.__average_salary_dict[key] = int(mean(value))
        return self.__average_salary_dict


class CountDict:
    """Класс статистики"""
    def __init__(self):
        """Инициализация объекта CountDict"""
        self.length = 0
        self.amount_dict = {}
        self.big_towns_arr = []
        self.prevailing_dict = {}

    def update_amount(self, key):
        """Обновить значения по ключу"""
        if self.amount_dict.get(key) is None:
            self.amount_dict[key] = 0
        self.amount_dict[key] += 1
        self.length += 1
        return

    def calculate_proportion(self):
        """Посчитать пропорцию"""
        proportion_dict = {}
        for key, value in self.amount_dict.items():
            proportion = value / self.length
            if proportion >= 0.1:
                self.big_towns_arr.append(key)
                proportion_dict[key] = round(proportion, 4)
        sorted_dict = dict(sorted(proportion_dict.items(), key=lambda row: row[1], reverse=True))
        self.prevailing_dict = {x: sorted_dict[x] for x in list(sorted_dict)[:10]}
        return


class Constractor:
    """Собрать полученные статистические данные в одну структуру"""
    def __init__(self):
        """Инициализация объекта Constractor"""
        self.year_salary = SalaryDict()
        self.year_vacancy_amount = CountDict()
        self.year_vacancy_salary = SalaryDict()
        self.year_position_vacancy_amount = CountDict()
        self.town_salary = SalaryDict()
        self.town_job_rating = CountDict()

    def compile_data(self, vacancies, prof):
        """Получить и объединить данные статистики из словаря

            Args:
                vacancies: вакансия
                prof (str): название профессии

            Returns: tuple: (средняя зарплата, годовое количество мест, средняя зарплата по вакансии, количество
            вакансий по годам, зарплата по городам, городской рейтинг профессий)
        """
        self.calculate_stat_values(prof, vacancies)
        if self.year_vacancy_salary.salary_dict == {}:
            self.year_vacancy_salary.salary_dict = {x: [0] for x in self.year_salary.salary_dict.keys()}
        elif self.year_vacancy_salary.salary_dict != {} and len(
                list(self.year_salary.calculate_average_salary().keys())) != len(
            list(self.year_vacancy_salary.calculate_average_salary().keys())):
            for key in list(self.year_salary.calculate_average_salary().keys()):
                self.set_key_to_zero(key)
        if self.year_position_vacancy_amount.amount_dict == {}:
            self.year_position_vacancy_amount.amount_dict = {x: 0 for x in self.year_vacancy_amount.amount_dict.keys()}
        elif self.year_position_vacancy_amount.amount_dict != {} and len(
                list(self.year_vacancy_amount.amount_dict.keys())) != len(
            list(self.year_position_vacancy_amount.amount_dict.keys())):
            for key in list(self.year_vacancy_amount.amount_dict.keys()):
                self.set_position_vacancy_amount_to_zero(key)
        self.town_salary, del_for_towns = self.calculate_highest_average_salary(self.town_salary)
        self.town_job_rating.calculate_proportion()
        self.town_job_rating = self.find_highest_town_rating(self.town_job_rating)
        self.town_job_rating = dict((x, y) for x, y in self.town_job_rating)
        return self.year_salary.calculate_average_salary(), self.year_vacancy_amount.amount_dict, \
               self.year_vacancy_salary.calculate_average_salary(), self.year_position_vacancy_amount.amount_dict, \
               self.town_salary, self.town_job_rating

    def set_position_vacancy_amount_to_zero(self, key):
        """Вспомогательная функция для приведения значения количества профессий к нулю по ключу

            Args:
                key: ключ

        """
        if key not in list(self.year_position_vacancy_amount.amount_dict.keys()):
            self.year_position_vacancy_amount.amount_dict[key] = 0

    def set_key_to_zero(self, key):
        """Вспомогательная функция для приведения значения количества профессий к нулю по ключу

            Args:
                key: ключ
        """
        if key not in list(self.year_vacancy_salary.calculate_average_salary().keys()):
            self.year_vacancy_salary.calculate_average_salary()[key] = 0

    def calculate_stat_values(self, prof, vacancies):
        """Обновить полученные значения пунктов

            Args:
                prof (str): название профессии
                vacancies: вакансии
        """
        for vacancy in vacancies:
            vacancy_salary = (vacancy.salary_from + vacancy.salary_to) / 2 * currency_to_rub[vacancy.salary_currency]
            self.year_salary.append_salary(vacancy.published_at, vacancy_salary)
            self.year_vacancy_amount.update_amount(vacancy.published_at)
            self.town_salary.append_salary(vacancy.area_name, vacancy_salary)
            self.town_job_rating.update_amount(vacancy.area_name)
            if prof in vacancy.name:
                self.year_vacancy_salary.append_salary(vacancy.published_at, vacancy_salary)
                self.year_position_vacancy_amount.update_amount(vacancy.published_at)

    @staticmethod
    def calculate_highest_average_salary(list_all_salary):
        """"Высчитать самую высокую среднюю зарплату

            Args:
                list_all_salary: выдержка по всем зарплатам

            Returns:
                tuple: (высшая зарплата, фильтр списка городов)
        """
        average_salary_values = []
        town_tracker = {}
        for i in range(len(list_all_salary.salary_dict)):
            town = list(list_all_salary.salary_dict)[i]
            town_tracker[town] = len(list(list_all_salary.salary_dict.values())[i])
            average = int(sum(list(list_all_salary.salary_dict.values())[i]) / 
                       len(list(list_all_salary.salary_dict.values())[i]))
            average_salary_values.append((town, average))

        del_for_towns = []
        del_town_index = []
        for i in range(len(town_tracker.items())):
            town = list(town_tracker)[i]
            percentage = round(100 * int(list(town_tracker.values())[i]) / sum(town_tracker.values()), 1)
            if percentage < 1 or town == 'Россия':
                del_for_towns.append((town, list(town_tracker.values())[i]))
                del_town_index.append(i)

        for i in reversed(range(len(del_for_towns))):
            del town_tracker[del_for_towns[i][0]]
            del average_salary_values[del_town_index[i]]

        top_aver_salary = dict(sorted(average_salary_values, key=lambda row: row[1], reverse=True))
        big_salary_dict = {}
        for key, value in top_aver_salary.items():
            big_salary_dict[key] = value
        return {x: big_salary_dict[x] for x in list(big_salary_dict)[:10]}, del_for_towns

    @staticmethod
    def find_highest_town_rating(town_job_rating):
        """Найти рейтинговую пропорцию профессии по городам

            Args:
                town_job_rating: рейтинг профессии по городам

            Returns:
                dict: отсортированный словарь пропорции
        """
        del_for_towns = []
        for i in reversed(range(len(del_for_towns))):
            del town_job_rating.amount_dict[del_for_towns[i][0]]
        for i in range(len(town_job_rating.amount_dict.keys())):
            if 'Россия' in town_job_rating.amount_dict.keys():
                del town_job_rating.amount_dict['Россия']

        proportion_dict = {}
        
        for key, value in town_job_rating.amount_dict.items():
            proportion = value / town_job_rating.length
            if proportion >= 0.01:
                proportion_dict[key] = round(proportion, 4)

        sorted_dict = sorted(proportion_dict.items(), key=lambda row: row[1], reverse=True)
        return sorted_dict[:10]


class CreateReport:
    """Класс для создания отчёта"""
    def __init__(self, data, prof):
        """Инициализация объекта CreateReport

            Args:
                data: полученные в ходе подсчётов данные
                prof (str): название профессии
        """
        self.year_salary = data[0]
        self.year_vacancy_amount = data[1]
        self.year_vacancy_salary = data[2]
        self.year_position_vacancy_amount = data[3]
        self.town_salary = data[4]
        self.town_job_rating = data[5]
        self.prof = prof

        self.wb = Workbook()
        self.sheet1 = self.wb.active
        self.sheet1.title = 'Статистика по годам'
        self.sheet2 = self.wb.create_sheet('Статистика по городам')

        self.fig = plt.figure()
        self.ax1 = self.fig.add_subplot(221)
        self.ax1.set_title('Уровень зарплат по годам')
        self.ax2 = self.fig.add_subplot(222)
        self.ax2.set_title('Количество вакансий по годам')
        self.ax3 = self.fig.add_subplot(223)
        self.ax3.set_title('Уровень зарплат по городам')
        self.ax4 = self.fig.add_subplot(224)
        self.ax4.set_title('Доля вакансий по городам')

    def create_excel_sheets(self):
        """Создать файл Excel"""
        names_sheet1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.prof}',
                        'Количество вакансий', f'Количество вакансий - {self.prof}']
        names_sheet2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']

        for i, name in enumerate(names_sheet1):
            self.sheet1.cell(row=1, column=(i + 1), value=name).font = Font(bold=True)
        for year, value in self.year_salary.items():
            self.sheet1.append([year, value, self.year_vacancy_salary[year], self.year_vacancy_amount[year],
                                self.year_position_vacancy_amount[year]])

        for i, name in enumerate(names_sheet2):
            self.sheet2.cell(row=1, column=(i + 1), value=name).font = Font(bold=True)
        for i in range(len(list(self.town_salary.keys()))):
            self.sheet2.append([list(self.town_salary.keys())[i], list(self.town_salary.values())[i],
                                list(self.town_job_rating.keys())[i], list(self.town_job_rating.values())[i]])

        cell_border = Side(border_style='thin', color='000000')
        self.set_border(self.sheet1, cell_border)
        self.set_border(self.sheet2, cell_border)
        self.sheet2.insert_cols(3)
        self.sheet2.column_dimensions['C'].width = 2

        self.calculate_column_width(self.sheet1)
        self.calculate_column_width(self.sheet2)

        for i in range(2, len(self.sheet2['E']) + 1):
            self.sheet2[f'E{i}'].number_format = FORMAT_PERCENTAGE_00

        self.wb.save('report.xlsx')

    @staticmethod
    def set_border(ws, cell_border):
        """Установить границы таблицы Excel

            Args:
                ws: рабочая плоскость
                cell_border (int): граница
        """
        for cell in ws._cells.values():
            cell.border = Border(top=cell_border, bottom=cell_border, left=cell_border, right=cell_border)

    @staticmethod
    def calculate_column_width(ws):
        """Посчитать ширину колонки

            Args:
                ws: рабочая плоскость
        """
        dimension_dict = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dimension_dict[cell.column_letter] = max((dimension_dict.get(cell.column_letter, 0), len(str(cell.value))))
        for column_var, value in dimension_dict.items():
            ws.column_dimensions[column_var].width = value + 2

    def create_image(self):
        """Создать картинку со статистическими данными"""
        width_12 = 0.4
        x_nums_1 = np.arange(len(self.year_salary.keys()))
        x_list1_1 = x_nums_1 - width_12 / 2
        x_list1_2 = x_nums_1 + width_12 / 2

        self.ax1.bar(x_list1_1, self.year_salary.values(), width_12, label='средняя з/п')
        self.ax1.bar(x_list1_2, self.year_vacancy_salary.values(), width_12, label=f'з/п {self.prof}')
        self.ax1.set_xticks(x_nums_1, self.year_salary.keys(), rotation='vertical')
        self.ax1.tick_params(axis='both', labelsize=8)
        self.ax1.legend(fontsize=8)
        self.ax1.grid(True, axis='y')

        x_nums_2 = np.arange(len(self.year_vacancy_amount.keys()))
        x_list2_1 = x_nums_2 - width_12 / 2
        x_list2_2 = x_nums_2 + width_12 / 2

        self.ax2.bar(x_list2_1, self.year_vacancy_amount.values(), width_12, label='Количество вакансий')
        self.ax2.bar(x_list2_2, self.year_position_vacancy_amount.values(), width_12, label=f'Количество вакансий\n{self.prof}')
        self.ax2.set_xticks(x_nums_2, self.year_vacancy_amount.keys(), rotation='vertical')
        self.ax2.tick_params(axis='both', labelsize=8)
        self.ax2.legend(fontsize=8)
        self.ax2.grid(True, axis='y')

        img_titles = {}
        self.apply_attributes(img_titles)

        width_3 = 0.7
        y_nums = np.arange(len(list(img_titles.keys())))

        self.ax3.barh(y_nums, img_titles.values(), width_3, align='center')
        self.ax3.set_yticks(y_nums, img_titles.keys())
        self.ax3.tick_params(axis='y', labelsize=6)
        self.ax3.tick_params(axis='x', labelsize=8)
        self.ax3.invert_yaxis()
        self.ax3.grid(True, axis='x')

        other = 1
        data = [1]
        labels = ['Другие']
        for key, value in self.town_job_rating.items():
            data.append(value * 100)
            labels.append(key)
            other -= value
        data[0] = round(other, 4) * 100
        textprops = {"fontsize": 6}

        self.ax4.pie(data, labels=labels, textprops=textprops, radius=1.1)

        plt.tight_layout()
        plt.savefig('graph.png')

    def apply_attributes(self, img_titles):
        """Задать параметры текста на картинке

            Args:
                img_titles (str): заголовки пунктов на картинках
        """
        for key, value in self.town_salary.items():
            if ' ' in key:
                key = str(key).replace(' ', '\n')
            elif '-' in key and key.count('-') == 1:
                key = str(key).replace('-', '-\n')
            elif '-' in key and key.count('-') != 1:
                key = str(key).replace('-', '-\n', 1)
            img_titles[key] = value

    def create_pdf(self):
        """Создать PDF-файл по всей статистике"""
        new_environment = Environment(loader=FileSystemLoader('.'))
        template = new_environment.get_template('pdf_template.html')
        names_sheet1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.prof}',
                        'Количество вакансий', f'Количество вакансий - {self.prof}']
        names_sheet2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']
        town_stats_compilation = []
        for i in range(len(list(self.town_salary.keys()))):
            town_stats_compilation.append([list(self.town_salary.keys())[i], list(self.town_job_rating.keys())[i]])
        for key, value in self.town_job_rating.items():
            self.town_job_rating[key] = str(round(value * 100, 2)) + '%'
        pdf_template = self.create_pdf_template(names_sheet1, names_sheet2, template, town_stats_compilation)
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

    def create_pdf_template(self, names_sheet1, names_sheet2, template, town_stats_compilation):
        """Создать шаблон


            Args:
                names_sheet1: лист статистики 1
                names_sheet2: лист статистики 2
                template: шаблон
                town_stats_compilation: статистика по городам
        """
        pdf_template = template.render({'name': self.prof,
                                        'year_salary': self.year_salary,
                                        'year_vacancy_amount': self.year_vacancy_amount,
                                        'year_vacancy_salary': self.year_vacancy_salary,
                                        'year_position_vacancy_amount': self.year_position_vacancy_amount,
                                        'town_stats_compilation': town_stats_compilation,
                                        'town_salary': self.town_salary,
                                        'town_job_rating': self.town_job_rating,
                                        'names_sheet1': names_sheet1,
                                        'names_sheet2': names_sheet2})
        return pdf_template


def generate_output(data_vacancies, position_title):
    """Сгенерировать вывод по шаблону

        Args:
            data_vacancies: данные вакансий
            position_title: название профессии
    """
    compiled_vacancies_arr = []
    for compilation in data_vacancies:
        compilation = Vacancy(dict(zip(column_headers, compilation)))
        compiled_vacancies_arr.append(compilation)
    data = Constractor()
    data = data.compile_data(compiled_vacancies_arr, position_title)

    print(f'Динамика уровня зарплат по годам: {data[0]}')
    print(f'Динамика количества вакансий по годам: {data[1]}')
    print(f'Динамика уровня зарплат по годам для выбранной профессии: {data[2]}')
    print(f'Динамика количества вакансий по годам для выбранной профессии: {data[3]}')
    print(f'Уровень зарплат по городам (в порядке убывания): {data[4]}')
    print(f'Доля вакансий по городам (в порядке убывания): {data[5]}')
    
    return data


users_input = UsersInput()
requested_data = DataSet(users_input.compiled_file)
column_headers, vacancies_data = requested_data.columns_names, requested_data.vacancies_data
output_data = generate_output(vacancies_data, users_input.position_title)
generated_report = CreateReport(output_data, users_input.position_title)
generated_report.create_excel_sheets()
generated_report.create_image()
generated_report.create_pdf()
